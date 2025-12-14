import psycopg2
from openpyxl import Workbook
from datetime import datetime


def export_logs(
    user="postgres",
    password='festopostgre',
    host="localhost",
    port=5432,
    exclude_system_dbs=True
):
    """
    Exports EVERY PostgreSQL database into its own .xlsx file.
    Each table inside each database becomes a worksheet.
    """

    print("[INFO] Connecting to PostgreSQL to list databases...")

    # Connect to the main postgres DB
    if password:
        conn = psycopg2.connect(dbname="postgres", user=user, password=password, host=host, port=port)
    else:
        conn = psycopg2.connect(dbname="postgres", user=user)

    cur = conn.cursor()

    # Get list of databases
    cur.execute("""
        SELECT datname 
        FROM pg_database
        WHERE datistemplate = false;
    """)

    dbs = [row[0] for row in cur.fetchall()]
    cur.close()
    conn.close()

    # Optionally remove system databases
    if exclude_system_dbs:
        for sysdb in ["postgres", "template1", "template0"]:
            if sysdb in dbs:
                dbs.remove(sysdb)

    print(f"[INFO] Found {len(dbs)} user databases:")
    for db in dbs:
        print(f"  - {db}")

    # Export each database
    for db_name in dbs:
        export_single_database_to_excel(
            db_name=db_name,
            user=user,
            password=password,
            host=host,
            port=port
        )


def export_single_database_to_excel(
    db_name,
    user="postgres",
    password=None,
    host="localhost",
    port=5432
):
    """
    Export one DB: each table → one sheet.
    Automatically strips timezone (tzinfo) from datetime values.
    """

    print(f"\n[INFO] Exporting database '{db_name}'...")

    # Connect
    if password:
        conn = psycopg2.connect(dbname=db_name, user=user, password=password, host=host, port=port)
    else:
        conn = psycopg2.connect(dbname=db_name, user=user)

    cur = conn.cursor()

    # Get tables
    cur.execute("""
        SELECT table_name
        FROM information_schema.tables
        WHERE table_schema = 'public'
        ORDER BY table_name;
    """)

    tables = [row[0] for row in cur.fetchall()]

    if not tables:
        print(f"[WARNING] No tables found in '{db_name}'. Skipping.")
        return

    # Create Excel workbook
    wb = Workbook()
    wb.remove(wb.active)

    for table in tables:
        print(f"[INFO]   Exporting table '{table}'")

        ws = wb.create_sheet(title=table[:31])  # Excel sheet name limit

        cur.execute(f"SELECT * FROM {table};")
        rows = cur.fetchall()
        columns = [desc[0] for desc in cur.description]

        ws.append(columns)

    for row in rows:
        cleaned_row = []

        for value in row:
            # Remove timezone information
            if hasattr(value, "tzinfo") and value.tzinfo is not None:
                value = value.replace(tzinfo=None)
            cleaned_row.append(value)

        ws.append(cleaned_row)

        # Apply millisecond format to datetime cells
        row_idx = ws.max_row
        for col_idx, value in enumerate(cleaned_row, start=1):
            if isinstance(value, datetime):
                ws.cell(row=row_idx, column=col_idx).number_format = "yyyy-mm-dd hh:mm:ss.000"

    # Save file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{db_name}_export_{timestamp}.xlsx"

    wb.save(filename)
    print(f"[SUCCESS] Database '{db_name}' exported to {filename}")

    cur.close()
    conn.close()